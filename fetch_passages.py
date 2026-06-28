#!/usr/bin/env python3
"""
지문 원문 추출 도구 (큰 시트 잘림 문제 해결)
===============================================
통합분석어시스트 시트의 '지문DB' 탭에서 모든 지문 원문을 ID→본문으로 추출.

⚠️ read_file_content는 977KB+ 시트를 잘라서 ABSO 11~16강 등이 누락됨.
   → 반드시 xlsx로 받아 openpyxl로 파싱해야 전체(230행)가 나온다.

[새 세션 사용법]
1) 시트를 xlsx로 다운로드 (Google Drive:download_file_content):
   fileId = '1_dt7nLh4c9eIRL-05zrGOHFc_m39baVZ7tCelAhRYzw'  (통합분석어시스트)
   exportMimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
2) tool_results의 base64 content를 디코드해 sheet.xlsx로 저장 (아래 decode_xlsx 참고)
3) python3 fetch_passages.py sheet.xlsx [ID접두어]
   예: python3 fetch_passages.py sheet.xlsx ABSO_TYPE
       python3 fetch_passages.py sheet.xlsx          (전체)
   → passages_extracted.json 에 {id: 본문} 저장

[base64 디코드 헬퍼] tool_results JSON에서 xlsx 만들기:
   import json,base64
   raw=json.load(open('<tool_result.json>'))
   txt=''.join(b['text'] for b in raw if isinstance(b,dict) and 'text' in b)
   obj=json.loads(txt); open('sheet.xlsx','wb').write(base64.b64decode(obj['content']))
"""
import sys, json, re

def extract(xlsx_path, prefix=''):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    if '지문DB' not in wb.sheetnames:
        print(f"⚠️ '지문DB' 탭 없음. 탭 목록: {wb.sheetnames}")
        return {}
    sh = wb['지문DB']
    rows = list(sh.iter_rows(values_only=True))
    hdr = rows[0]
    id_i = hdr.index('id') if 'id' in hdr else 0
    sent_i = hdr.index('sentences_json') if 'sentences_json' in hdr else 7
    out = {}
    for r in rows[1:]:
        rid = r[id_i]
        if not rid: continue
        rid = str(rid)
        if prefix and not rid.startswith(prefix): continue
        sj = r[sent_i]
        if not sj: continue
        try:
            sents = json.loads(sj)
            body = ' '.join(s['text'] for s in sents if isinstance(s, dict) and 'text' in s)
            if body: out[rid] = body
        except Exception:
            pass
    return out

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(2)
    xlsx = sys.argv[1]
    prefix = sys.argv[2] if len(sys.argv) > 2 else ''
    passages = extract(xlsx, prefix)
    json.dump(passages, open('passages_extracted.json', 'w'), ensure_ascii=False)
    # 교재·강별 요약
    chaps = {}
    for pid in passages:
        m = re.match(r'([A-Z_]+?)_?(\d+)_\d+$', pid) or re.match(r'(\D+)(\d+)', pid)
        book = re.split(r'_\d', pid)[0]
        chaps.setdefault(book, set()).add(re.search(r'_(\d+)_\d+$', pid).group(1) if re.search(r'_(\d+)_\d+$', pid) else '?')
    print(f"✅ 추출 완료: {len(passages)}개 지문 → passages_extracted.json")
    for book, cs in sorted(chaps.items()):
        print(f"   {book}: {len(cs)}개 강 ({','.join(sorted(cs))})")
